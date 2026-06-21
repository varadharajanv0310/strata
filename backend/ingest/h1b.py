"""Real **US disclosed wages** from DOL OFLC LCA (H-1B) disclosure data.

Every H-1B petition files a Labor Condition Application whose *offered wage* the
DOL publishes quarterly as a large LCA disclosure spreadsheet. That is a census
of real, employer-disclosed pay (not survey, not advertised) — person-level wage
records keyed by SOC occupation, worksite and case status. We stream each file
with ``openpyxl`` read-only (the files are 80-250 MB — never load whole-file in
pandas), keep only ``Certified`` cases, annualize the offered wage, crosswalk the
SOC code to our 16 role ids, and aggregate to **median annual USD wage per
(role, US, pooled, year)**.

Source: ``https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_FY<yyyy>_Q<n>.xlsx``.
The dol.gov edge (Akamai) 403s non-browser clients, so we fetch the *identical
bytes* from the Internet Archive's raw mirror (``id_`` snapshot) — same file,
verified ``PK`` zip / openxml content-type, and it honours HTTP range requests so
the multi-hundred-MB download resumes instead of restarting (brief §10).

Cache (the per-quarter ``.xlsx`` + the aggregate JSON) is the checkpoint —
idempotent + resumable. Credential-free public data; network-resilient with
bounded retry then give-up-and-continue per unit.
"""
from __future__ import annotations

import json
import os
import statistics
import time
from collections import defaultdict

import requests

from backend.core.config import settings
from backend.core.logging import get_logger

log = get_logger("ingest.h1b")

# Canonical DOL file (the real, citable URL) and the Internet-Archive raw-byte
# mirror we actually fetch (dol.gov 403s datacenter traffic). ``{ts}id_`` returns
# the original archived bytes unmodified. Snapshots verified 2026-06.
DOL_CANONICAL = ("https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/"
                 "LCA_Disclosure_Data_{fy}_{q}.xlsx")
WAYBACK_RAW = ("https://web.archive.org/web/{ts}id_/https://www.dol.gov/sites/"
               "dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_{fy}_{q}.xlsx")

# known-good Internet-Archive snapshot timestamps per (fy, quarter)
_SNAPSHOTS = {
    ("FY2025", "Q4"): "20260114000337",
    ("FY2025", "Q3"): "20260113235849",
    ("FY2025", "Q2"): "20260114000118",
    ("FY2025", "Q1"): "20260113235745",
    ("FY2024", "Q4"): "20260114000455",
    ("FY2023", "Q4"): "20260113235610",
}
# the file's fiscal year → calendar "year" we tag the aggregate with
_FY_YEAR = {"FY2025": 2025, "FY2024": 2024, "FY2023": 2023, "FY2022": 2022, "FY2021": 2021}

_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
_HDRS = {"User-Agent": _UA, "Accept": "*/*"}

# SOC crosswalk → our 16 role ids.  H-1B LCA codes the occupation in the
# **8-digit O*NET-SOC** form (e.g. ``15-1252.00``, ``15-1299.08``); the trailing
# ``.NN`` distinguishes occupations that share a 6-digit SOC base — e.g. under
# ``15-1299`` (Computer Occupations, All Other) sit *Computer Systems Engineers*
# (.08 → backend), *IT Project Managers* (.09 → pm) and *Information Security
# Engineers* (.05 → security).  The previous parser truncated to 7 chars BEFORE
# the lookup, collapsing every ``15-1299.xx`` to a single ``swe`` and losing six
# of our roles entirely.  We now match the FULL O*NET code first (``SOC_ONET``),
# then fall back to the 6-digit base (``SOC_BASE``), which also catches the older
# legacy 2010-SOC codes and the rows the file stores already-truncated.

# --- exact 8-digit O*NET-SOC overrides (most specific; checked first) ---
SOC_ONET = {
    "15-1299.08": "backend",     # Computer Systems Engineers/Architects
    "15-1299.09": "pm",          # Information Technology Project Managers
    "15-1299.05": "security",    # Information Security Engineers
    "15-1299.07": "backend",     # Network/Systems Engineers (legacy O*NET) -> backend
    "15-2051.01": "data-analyst", # Business Intelligence Analysts
    "15-2051.02": "data-sci",    # Clinical Data Managers -> data-sci
    "15-1243.01": "data-eng",    # Data Warehousing Specialists
    "15-2041.01": "data-sci",    # Biostatisticians -> data-sci
    "15-1255.00": "ux",          # Web & Digital Interface Designers -> ux/design
}

# --- 6-digit SOC base (2018) + legacy 2010 codes → role ---
SOC_BASE = {
    # ---- software / general engineering ----
    "15-1252": "swe",          # Software Developers
    "15-1132": "swe",          # Software Developers, Applications (2010)
    "15-1133": "swe",          # Software Developers, Systems Software (2010)
    "15-1251": "swe",          # Computer Programmers
    "15-1131": "swe",          # Computer Programmers (2010)
    "15-1211": "swe",          # Computer Systems Analysts
    "15-1121": "swe",          # Computer Systems Analysts (2010)
    "15-1199": "swe",          # Computer Occupations, All Other (2010)
    "15-1299": "swe",          # Computer Occupations, All Other (base fallback)
    "17-2061": "swe",          # Computer Hardware Engineers (hardware-adjacent)
    # ---- ml / research ----
    "15-1221": "ml-eng",       # Computer & Information Research Scientists
    "15-2098": "data-sci",     # Data Scientists & Math Science (provisional, 2018→)
    # ---- frontend / web ----
    "15-1254": "frontend",     # Web Developers
    "15-1134": "frontend",     # Web Developers (2010)
    "15-1255": "frontend",     # Web & Digital Interface Designers (base → frontend)
    # ---- ux / design ----
    "27-1024": "ux",           # Graphic Designers
    "27-1021": "ux",           # Commercial & Industrial Designers
    "27-1014": "ux",           # Multimedia Artists & Animators
    # ---- qa ----
    "15-1253": "qa",           # Software QA Analysts & Testers
    "15-1232": "qa",           # Computer User Support Specialists -> qa/IT
    "15-1151": "qa",           # Computer User Support Specialists (2010)
    # ---- data science / analytics ----
    "15-2051": "data-sci",     # Data Scientists
    "15-2031": "data-sci",     # Operations Research Analysts
    "15-2041": "data-analyst", # Statisticians
    "13-2099": "data-analyst", # Financial Specialists, All Other (often analyst)
    "13-1161": "data-analyst", # Market Research Analysts
    "13-1111": "data-analyst", # Management Analysts
    # ---- data engineering ----
    "15-1243": "data-eng",     # Database & Network Architects
    "15-1242": "data-eng",     # Database Administrators
    "15-1141": "data-eng",     # Database Administrators (2010)
    "15-1245": "data-eng",     # Database Admins & Architects
    # ---- security ----
    "15-1212": "security",     # Information Security Analysts
    "15-1122": "security",     # Information Security Analysts (2010)
    # ---- cloud architecture ----
    "15-1241": "cloud-arch",   # Computer Network Architects
    "15-1143": "cloud-arch",   # Computer Network Architects (2010)
    # ---- devops / infra ----
    "15-1244": "devops",       # Network & Computer Systems Administrators
    "15-1142": "devops",       # Network & Computer Systems Administrators (2010)
    # ---- product management ----
    "13-1082": "pm",           # Project Management Specialists
    "11-3061": "pm",           # Purchasing Managers (product ops adjacent) -- skip below
    # ---- engineering management ----
    "11-3021": "eng-mgr",      # Computer & Information Systems Managers
    "11-9041": "eng-mgr",      # Architectural & Engineering Managers
}
# trim a borderline mapping we listed for documentation but don't want counted
SOC_BASE.pop("11-3061", None)


def _soc_to_role(raw) -> str | None:
    """Map a raw SOC/O*NET cell to a role id: full O*NET code first, then base."""
    s = str(raw or "").strip()
    if not s:
        return None
    # normalize: some files store '15-1252' some '15-1252.00' some '151252'
    if len(s) == 6 and s.isdigit():           # e.g. '151252' -> '15-1252'
        s = s[:2] + "-" + s[2:]
    rid = SOC_ONET.get(s)                      # exact 8-digit O*NET override
    if rid:
        return rid
    base = s[:7] if len(s) >= 7 else s         # '15-1299.08' -> '15-1299'
    return SOC_BASE.get(base)

# annualize an offered wage to a yearly figure
_UNIT_FACTOR = {
    "year": 1.0, "yr": 1.0, "annual": 1.0,
    "hour": 2080.0, "hr": 2080.0, "hourly": 2080.0,
    "month": 12.0, "mth": 12.0, "monthly": 12.0,
    "week": 52.0, "wk": 52.0, "weekly": 52.0,
    "bi-weekly": 26.0, "biweekly": 26.0, "bi weekly": 26.0,
}
WAGE_LO, WAGE_HI = 20_000.0, 1_000_000.0  # trim garbage / data-entry errors

# header name candidates (vary slightly by FY)
_C_STATUS = ["CASE_STATUS"]
_C_SOC = ["SOC_CODE", "SOC_CODE_1", "OCCUPATIONAL_CODE", "OES_SOC_CODE"]
_C_WAGE_FROM = ["WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_FROM_1", "WAGE_RATE_OF_PAY"]
_C_WAGE_TO = ["WAGE_RATE_OF_PAY_TO", "WAGE_RATE_OF_PAY_TO_1"]
_C_UNIT = ["WAGE_UNIT_OF_PAY", "WAGE_UNIT_OF_PAY_1", "PW_UNIT_OF_PAY"]
_C_STATE = ["WORKSITE_STATE", "WORKSITE_STATE_1", "EMPLOYER_STATE", "WORKLOC1_STATE"]
_C_LEVEL = ["PW_WAGE_LEVEL", "PW_WAGE_LEVEL_1", "WAGE_LEVEL"]

# DOL prevailing-wage levels map to BLS experience tiers (20 CFR 656.40 guidance:
# Level I = entry / <2 yr, II = qualified, III = experienced, IV = fully
# competent / senior).  We bucket each Certified filing into one of our warehouse
# experience bands AND, separately, into the pooled band.
_LEVEL_TO_BAND = {
    "i": "0-2", "1": "0-2", "level i": "0-2", "entry": "0-2",
    "ii": "3-5", "2": "3-5", "level ii": "3-5",
    "iii": "6-9", "3": "6-9", "level iii": "6-9",
    "iv": "10+", "4": "10+", "level iv": "10+",
}


def _staging_dir():
    d = settings.staging_dir / "h1b"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _file_path(fy: str, q: str):
    return _staging_dir() / f"LCA_{fy}_{q}.xlsx"


def _expected_size(fy: str, q: str) -> int | None:
    ts = _SNAPSHOTS.get((fy, q))
    if not ts:
        return None
    url = WAYBACK_RAW.format(ts=ts, fy=fy, q=q)
    try:
        r = requests.head(url, headers=_HDRS, timeout=60, allow_redirects=True)
        cl = r.headers.get("Content-Length")
        return int(cl) if cl else None
    except Exception:  # noqa: BLE001
        return None


def _download(fy: str, q: str, max_retries: int = 6) -> bool:
    """Resumable range-download of one quarter's xlsx. Cache file IS the checkpoint."""
    ts = _SNAPSHOTS.get((fy, q))
    if not ts:
        log.error("h1b %s %s: no known snapshot — skipping", fy, q)
        return False
    p = _file_path(fy, q)
    url = WAYBACK_RAW.format(ts=ts, fy=fy, q=q)
    expected = _expected_size(fy, q)
    if p.exists() and expected and p.stat().st_size == expected:
        log.info("h1b %s %s already cached (%d bytes)", fy, q, expected)
        return True

    for attempt in range(max_retries):
        have = p.stat().st_size if p.exists() else 0
        if expected and have == expected:
            return True
        if expected and have > expected:  # corrupt/partial mismatch — restart clean
            p.unlink(missing_ok=True)
            have = 0
        headers = dict(_HDRS)
        mode = "wb"
        if have > 0:
            headers["Range"] = f"bytes={have}-"
            mode = "ab"
        try:
            with requests.get(url, headers=headers, timeout=180, stream=True) as r:
                if have > 0 and r.status_code == 200:  # server ignored range — restart
                    p.unlink(missing_ok=True)
                    mode, have = "wb", 0
                elif r.status_code not in (200, 206):
                    r.raise_for_status()
                with open(p, mode) as f:
                    for chunk in r.iter_content(1024 * 256):
                        if chunk:
                            f.write(chunk)
            got = p.stat().st_size
            if not expected or got >= expected:
                log.info("h1b %s %s downloaded (%d bytes)", fy, q, got)
                return True
            log.warning("h1b %s %s short (%d/%d) — resuming", fy, q, got, expected)
        except Exception as e:  # noqa: BLE001 — resilient: backoff then resume
            wait = min(60, 4 * (2 ** attempt))
            log.warning("h1b %s %s download error (try %d): %s — backoff %ss",
                        fy, q, attempt + 1, e, wait)
            time.sleep(wait)
    log.error("h1b %s %s: gave up after %d retries", fy, q, max_retries)
    return False


def _pick(header_map: dict, candidates) -> int | None:
    for c in candidates:
        if c in header_map:
            return header_map[c]
    return None


def _annualize(raw, unit) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    if val <= 0:
        return None
    u = (str(unit or "")).strip().lower()
    factor = _UNIT_FACTOR.get(u)
    if factor is None:  # default-assume yearly only when the number is plausibly annual
        factor = 1.0 if val >= 20_000 else None
    if factor is None:
        return None
    return val * factor


def _parse_file(fy: str, q: str, buckets: dict, max_rows: int | None = None,
                time_cap_s: float | None = None, start_t: float | None = None) -> int:
    """Stream the xlsx read-only; bucket annualized wages by (role, year, band).

    Every kept row is added to BOTH its experience band (from PW_WAGE_LEVEL) and
    the ``pooled`` band, so a role with filings yields a pooled cell plus up to
    four banded cells.  Prints a flushing heartbeat every 50k scanned rows.
    """
    import openpyxl  # local import: heavy, only needed when parsing

    p = _file_path(fy, q)
    year = _FY_YEAR.get(fy, int(fy[2:]))
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return 0
    hmap = {str(h).strip().upper(): i for i, h in enumerate(header) if h is not None}
    i_status = _pick(hmap, _C_STATUS)
    i_soc = _pick(hmap, _C_SOC)
    i_from = _pick(hmap, _C_WAGE_FROM)
    i_to = _pick(hmap, _C_WAGE_TO)
    i_unit = _pick(hmap, _C_UNIT)
    i_state = _pick(hmap, _C_STATE)
    i_level = _pick(hmap, _C_LEVEL)
    if i_soc is None or i_from is None or i_unit is None:
        log.error("h1b %s %s: missing key cols (soc=%s from=%s unit=%s) headers=%s",
                  fy, q, i_soc, i_from, i_unit, list(hmap)[:12])
        wb.close()
        return 0

    kept = seen = 0
    t0 = start_t if start_t is not None else time.time()
    for row in rows:
        seen += 1
        if max_rows is not None and seen > max_rows:
            break
        if seen % 50_000 == 0:  # streaming heartbeat
            el = time.time() - t0
            print(f"[h1b] {fy} {q}: scanned {seen:,} kept {kept:,} "
                  f"roles={len({k[0] for k in buckets})} elapsed {el:.0f}s",
                  flush=True)
            if time_cap_s is not None and el > time_cap_s:
                print(f"[h1b] {fy} {q}: time cap {time_cap_s}s hit — partial", flush=True)
                break
        if i_status is not None:
            st = str(row[i_status] or "").strip().lower()
            if not st.startswith("certified"):  # keep Certified / Certified-Withdrawn
                continue
        rid = _soc_to_role(row[i_soc])
        if not rid:
            continue
        wfrom = _annualize(row[i_from], row[i_unit])
        wto = _annualize(row[i_to], row[i_unit]) if i_to is not None else None
        if wfrom is None and wto is None:
            continue
        # use the FROM (offered floor); if only TO present, use it
        wage = wfrom if wfrom is not None else wto
        if wfrom is not None and wto is not None and wto >= wfrom:
            wage = (wfrom + wto) / 2.0  # midpoint of the disclosed band
        if not (WAGE_LO <= wage <= WAGE_HI):
            continue
        buckets[(rid, year, "pooled")].append(wage)
        band = _LEVEL_TO_BAND.get(str(row[i_level] or "").strip().lower()) if i_level is not None else None
        if band:
            buckets[(rid, year, band)].append(wage)
        kept += 1
    wb.close()
    el = time.time() - t0
    print(f"[h1b] {fy} {q} DONE: scanned {seen:,} kept {kept:,} "
          f"roles={len({k[0] for k in buckets})} elapsed {el:.0f}s", flush=True)
    log.info("h1b %s %s parsed: %d rows scanned, %d kept", fy, q, seen, kept)
    return kept


def _aggregate(buckets: dict, min_sample: int = 5) -> list:
    records = []
    for (rid, year, exp), vals in sorted(buckets.items()):
        if len(vals) < min_sample:
            continue
        n = len(vals)
        conf = "high" if n >= 500 else "med" if n >= 50 else "low"
        records.append({
            "role_id": rid, "country_code": "US", "experience_code": exp,
            "year": year, "median": round(statistics.median(vals)),
            "currency_code": "USD", "sample_size": n, "confidence": conf,
            "kind": "person-level", "source": "DOL OFLC H-1B/PERM",
        })
    return records


def _checkpoint(buckets: dict, out, min_sample: int = 5) -> int:
    """Write the current aggregate to disk (incremental checkpoint per file)."""
    records = _aggregate(buckets, min_sample=min_sample)
    out.write_text(json.dumps(records), encoding="utf-8")
    return len(records)


def run(years=None, quarters=None, max_rows=None, time_cap_s=None, min_sample=5) -> dict:
    """Download + parse DOL LCA disclosure quarters → median annual USD wage per
    (role, US, experience-band, year).

    Orchestrator entrypoint (full scale).

    Args:
      years: list of fiscal-year strings like ["FY2025","FY2024","FY2023"].
             Default = the most recent year we have a snapshot for.
      quarters: list like ["Q1","Q2","Q3","Q4"]. Default = all available for each year.
      max_rows: optional per-file row cap (used by the smoke test for speed).
      time_cap_s: wall-clock budget; the parse self-terminates (returns partial)
             once exceeded, checkpointing what it has.
      min_sample: minimum filings per cell to emit (banded cells are smaller).

    Checkpoints staging/h1b/salary_agg.json after EACH file and returns a summary.
    """
    if years is None:
        years = ["FY2025"]
    buckets: dict[tuple, list[float]] = defaultdict(list)
    out = _staging_dir() / "salary_agg.json"
    parsed_files, kept_total = [], 0
    start_t = time.time()
    for fy in years:
        qs = quarters or [q for (f, q) in _SNAPSHOTS if f == fy] or ["Q4", "Q3", "Q2", "Q1"]
        for q in qs:
            if (fy, q) not in _SNAPSHOTS:
                continue
            if time_cap_s is not None and (time.time() - start_t) > time_cap_s:
                print(f"[h1b] global time cap {time_cap_s}s hit before {fy} {q}", flush=True)
                break
            if not _download(fy, q):
                continue
            try:
                kept = _parse_file(fy, q, buckets, max_rows=max_rows,
                                   time_cap_s=time_cap_s, start_t=start_t)
            except Exception as e:  # noqa: BLE001 — bad file shouldn't kill the run
                log.error("h1b %s %s parse failed: %s", fy, q, e)
                continue
            kept_total += kept
            parsed_files.append(f"{fy}_{q}")
            cells = _checkpoint(buckets, out, min_sample=min_sample)  # incremental
            print(f"[h1b] checkpoint after {fy} {q}: {cells} cells, "
                  f"{kept_total:,} rows kept", flush=True)

    records = _aggregate(buckets, min_sample=min_sample)
    out.write_text(json.dumps(records), encoding="utf-8")
    by_role = defaultdict(int)
    for r in records:
        by_role[r["role_id"]] += 1
    summary = {
        "files_parsed": parsed_files,
        "rows_kept": kept_total,
        "cells": len(records),
        "roles": sorted({r["role_id"] for r in records}),
        "roles_count": len({r["role_id"] for r in records}),
        "cells_by_role": dict(sorted(by_role.items())),
        "out": str(out),
    }
    log.info("h1b aggregated: %s", summary)
    return summary


def smoke(max_rows: int | None = None, time_cap_s=None) -> dict:
    """Small real smoke test: parse the most recent fetchable quarter (FY2025 Q4)."""
    return run(years=["FY2025"], quarters=["Q4"], max_rows=max_rows, time_cap_s=time_cap_s)


def load_agg() -> list:
    p = _staging_dir() / "salary_agg.json"
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else []


if __name__ == "__main__":  # pragma: no cover
    print(json.dumps(smoke(), indent=2))
