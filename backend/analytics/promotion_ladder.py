"""Role promotion ladder — the dollar value of moving up a level in a ROLE.

strata is roles-only, so this answers a role question, not a company one: *if I move
up a seniority level within this role, what's the typical pay jump?* The DOL LCA
disclosure files we already hold carry, per certified filing, the **prevailing-wage
level (I–IV)** — a legally-attested seniority tier — the SOC occupation code, and the
offered wage. We crosswalk SOC → our role taxonomy, **pool every employer together**
(employers are aggregated away and never named or surfaced), and report the median
wage *step* between adjacent levels per role:

    Software Engineer (US)
      I  (entry)        $112k   n=41,203
      II (qualified)    $138k   n=88,510   ▲ +$26k (+23%)
      III(experienced)  $165k   n=52,330   ▲ +$27k (+20%)
      IV (senior)       $198k   n=14,880   ▲ +$33k (+20%)

(US-only — H-1B is the only at-scale person-level wage-by-level data; honest about
that.) Pure derivation on cached files — NO download, NO ingestion run. Reuses the
H-1B connector's SOC→role crosswalk + wage-annualization + level mapping; reads only
LCA xlsx already in ``staging/h1b``; empty result if none cached (never fetches).
"""
from __future__ import annotations

import statistics
import time
from collections import defaultdict

from backend.core.logging import get_logger
from backend.ingest.h1b import (
    _C_LEVEL,
    _C_SOC,
    _C_STATUS,
    _C_UNIT,
    _C_WAGE_FROM,
    _SNAPSHOTS,
    _annualize,
    _file_path,
    _pick,
    _soc_to_role,
    WAGE_HI,
    WAGE_LO,
)

log = get_logger("analytics.promotion_ladder")

_LEVEL_RANK = {"i": 1, "1": 1, "level i": 1, "ii": 2, "2": 2, "level ii": 2,
               "iii": 3, "3": 3, "level iii": 3, "iv": 4, "4": 4, "level iv": 4}
_LEVEL_LABEL = {1: "I (entry)", 2: "II (qualified)", 3: "III (experienced)", 4: "IV (senior)"}

# our role ids → display names (the SOC crosswalk in h1b maps to these)
ROLE_NAMES = {
    "swe": "Software Engineer", "backend": "Backend Engineer", "frontend": "Frontend Engineer",
    "mobile": "Mobile Engineer", "data-eng": "Data Engineer", "data-sci": "Data Scientist",
    "data-analyst": "Data Analyst", "ml-eng": "Machine Learning Engineer",
    "devops": "DevOps Engineer", "sre": "Site Reliability Engineer", "cloud-arch": "Cloud Architect",
    "security": "Security Engineer", "qa": "QA Engineer", "eng-mgr": "Engineering Manager",
    "pm": "Product Manager", "ux": "UX Designer",
}


def _cached_files() -> list[tuple[str, str]]:
    """(fy, q) pairs whose xlsx is actually cached in staging — newest first."""
    have = [(fy, q) for (fy, q) in _SNAPSHOTS if _file_path(fy, q).exists()]
    return sorted(have, reverse=True)


def build_role_ladders(
    files: list[tuple[str, str]] | None = None,
    *,
    max_rows: int | None = 600_000,
    min_rung_n: int = 30,
    min_levels: int = 2,
    time_cap_s: float = 180.0,
) -> dict:
    """Parse cached LCA xlsx → per-ROLE wage ladders (employers pooled, never named).

    Returns {"ladders": [...], "files": [...], "n_filings": int}. Each ladder:
    {role_id, role, n_filings, rungs:[{level,label,median,n}], steps:[{from,to,abs,pct}]}.
    """
    import openpyxl  # heavy; only on demand

    files = files or _cached_files()
    if not files:
        log.warning("no cached LCA xlsx in staging/h1b — run the H-1B connector first")
        return {"ladders": [], "files": [], "n_filings": 0}

    # (role_id, level_rank) -> [annualized wages]  — pooled across ALL employers
    buckets: dict[tuple[str, int], list[float]] = defaultdict(list)
    n_filings = 0
    t0 = time.time()

    for fy, q in files:
        p = _file_path(fy, q)
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if header is None:
            wb.close()
            continue
        hmap = {str(h).strip().upper(): i for i, h in enumerate(header) if h is not None}
        i_soc = _pick(hmap, _C_SOC)
        i_status = _pick(hmap, _C_STATUS)
        i_from = _pick(hmap, _C_WAGE_FROM)
        i_unit = _pick(hmap, _C_UNIT)
        i_level = _pick(hmap, _C_LEVEL)
        if None in (i_soc, i_from, i_unit, i_level):
            log.error("%s %s: missing cols (soc=%s from=%s unit=%s level=%s)",
                      fy, q, i_soc, i_from, i_unit, i_level)
            wb.close()
            continue

        seen = 0
        for row in it:
            seen += 1
            if max_rows is not None and seen > max_rows:
                break
            if seen % 50_000 == 0:
                el = time.time() - t0
                print(f"[ladder] {fy} {q}: scanned {seen:,} kept {n_filings:,} "
                      f"roles={len({k[0] for k in buckets})} elapsed {el:.0f}s", flush=True)
                if time_cap_s and el > time_cap_s:
                    print(f"[ladder] time cap {time_cap_s}s — partial", flush=True)
                    break
            if i_status is not None and not str(row[i_status] or "").strip().lower().startswith("certified"):
                continue
            rank = _LEVEL_RANK.get(str(row[i_level] or "").strip().lower())
            if not rank:
                continue
            role = _soc_to_role(row[i_soc])          # SOC → our role taxonomy (employer ignored)
            if not role:
                continue
            wage = _annualize(row[i_from], row[i_unit])
            if wage is None or not (WAGE_LO <= wage <= WAGE_HI):
                continue
            buckets[(role, rank)].append(wage)
            n_filings += 1
        wb.close()
        if time_cap_s and (time.time() - t0) > time_cap_s:
            break

    # assemble per-role ladders (a role needs >= min_levels populated rungs)
    grouped: dict[str, dict[int, list[float]]] = defaultdict(dict)
    for (role, rank), vals in buckets.items():
        if len(vals) >= min_rung_n:
            grouped[role][rank] = vals

    ladders = []
    for role, rungs in grouped.items():
        if len(rungs) < min_levels:
            continue
        ordered = sorted(rungs.items())
        rung_rows = [{"level": r, "label": _LEVEL_LABEL[r],
                      "median": round(statistics.median(v)), "n": len(v)}
                     for r, v in ordered]
        steps = []
        for a, b in zip(rung_rows, rung_rows[1:]):
            d = b["median"] - a["median"]
            steps.append({"from": a["label"], "to": b["label"], "abs": d,
                          "pct": round(100 * d / a["median"], 1) if a["median"] else 0.0})
        ladders.append({
            "role_id": role,
            "role": ROLE_NAMES.get(role, role),
            "country": "US",                        # H-1B is US-only
            "n_filings": sum(r["n"] for r in rung_rows),
            "rungs": rung_rows,
            "steps": steps,
            "total_climb_pct": round(100 * (rung_rows[-1]["median"] - rung_rows[0]["median"])
                                     / rung_rows[0]["median"], 1) if rung_rows[0]["median"] else 0.0,
        })
    ladders.sort(key=lambda L: L["n_filings"], reverse=True)
    log.info("role promotion ladders: %d filings → %d role ladders (>=%d levels) from %s",
             n_filings, len(ladders), min_levels, files)
    return {"ladders": ladders, "files": files, "n_filings": n_filings}


def _staging_file():
    from backend.core.config import settings
    d = settings.staging_dir / "analytics"
    d.mkdir(parents=True, exist_ok=True)
    return d / "role_ladders.json"


def run(**kw) -> dict:
    """Build the role pay ladders from cached H-1B xlsx → staging json (the serving
    materialize reads it). Pure derivation on cached files — no download/ingestion.
    Connector entrypoint (registered as a collect_all stage)."""
    import json
    res = build_role_ladders(**kw)
    _staging_file().write_text(json.dumps(res["ladders"]), encoding="utf-8")
    log.info("role pay ladders → staging: %d ladders from %d filings",
             len(res["ladders"]), res["n_filings"])
    return {"ladders": len(res["ladders"]), "n_filings": res["n_filings"], "written": True}


def load_ladders() -> list[dict]:
    import json
    f = _staging_file()
    return json.loads(f.read_text(encoding="utf-8")) if f.exists() else []


def sample(top: int = 16, **kw) -> dict:
    """Build role ladders and pretty-print them for a quick proof."""
    res = build_role_ladders(**kw)
    print(f"\n=== ROLE PROMOTION LADDERS (US · H-1B) "
          f"({res['n_filings']:,} filings → {len(res['ladders'])} roles, employers pooled) ===")
    for L in res["ladders"][:top]:
        print(f"\n{L['role']} ({L['country']})  (n={L['n_filings']:,}, climb +{L['total_climb_pct']}%)")
        for r in L["rungs"]:
            step = ""
            for s in L["steps"]:
                if s["to"] == r["label"]:
                    step = f"   ▲ +${s['abs']:,} (+{s['pct']}%)"
            print(f"   {r['label']:18} ${r['median']:>8,}  n={r['n']:<7,}{step}")
    return res


if __name__ == "__main__":  # pragma: no cover
    sample()
